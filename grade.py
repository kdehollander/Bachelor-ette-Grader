from ctypes import cast
from importlib.resources import contents
from tabnanny import check
import tkinter as tk
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import pickle

ROSE_CEREMONY_META = {
    "2nd Rose Ceremony": {
        "column": "D",
        "starting_cell": 7,
        "ending_cell": 24,
        "num_contestants": 18,
        "pp_correct_answer": 2,
        "weekly_score_cell": 27
    }, 
    "3rd Rose Ceremony": {
        "column": "F",
        "starting_cell": 7,
        "ending_cell": 21,
        "num_contestants": 15,
        "pp_correct_answer": 3,
        "weekly_score_cell": 27
    }, 
    "4th Rose Ceremony": {
        "column": "H",
        "starting_cell": 7,
        "ending_cell": 18,
        "num_contestants": 12,
        "pp_correct_answer": 4,
        "weekly_score_cell": 27
    }, 
    "5th Rose Ceremony": {
        "column": "J",
        "starting_cell": 7,
        "ending_cell": 15,
        "num_contestants": 9,
        "pp_correct_answer": 5,
        "weekly_score_cell": 27
    }, 
    "6th Rose Ceremony": {
        "column": "L",
        "starting_cell": 7,
        "ending_cell": 12,
        "num_contestants": 6,
        "pp_correct_answer": 6,
        "weekly_score_cell": 27
    }, 
    "7th Rose Ceremony": {
        "column": "N",
        "starting_cell": 7,
        "ending_cell": 10,
        "num_contestants": 4,
        "pp_correct_answer": 10,
        "weekly_score_cell": 27
    }, 
    "8th Rose Ceremony": {
        "column": "P",
        "starting_cell": 7,
        "ending_cell": 9,
        "num_contestants": 3,
        "pp_correct_answer": 15,
        "weekly_score_cell": 27
    }, 
    "9th Rose Ceremony": {
        "column": "R",
        "starting_cell": 7,
        "ending_cell": 8,
        "num_contestants": 2,
        "pp_correct_answer": 20,
        "weekly_score_cell": 27
    }, 
    "Final Rose": {
        "column": "T",
        "starting_cell": 7,
        "ending_cell": 7,
        "num_contestants": 1,
        "pp_correct_answer": 30,
        "weekly_score_cell": 27
    }
}

CONTESTANT_LIST = []
"""
CONTESTANT_LIST = ["Cassidy", "Claire", "Daria", "Eliza", "Elizabeth", "Ency", "Gabby", "Genevieve", "Hailey", "Hunter", "Ivana", "Jane", "Jill", "Kate", "Kira", "Lindsay D. ", "Lyndsey W. ", "Mara", "Marlena", "Melina", "Rachel", "Rianna", "Salley", "Samantha", "Sarah", "Serene", "Shanae", "Sierra", "Susie", "Teddi", "Tessa"]

with open('lib/contestant_list.pkl', 'wb') as f:
    pickle.dump(CONTESTANT_LIST, f)
"""
with open('lib/contestant_list.pkl', 'rb') as f:
    CONTESTANT_LIST = pickle.load(f)

def get_cast():
    bach_cast_url = "https://abc.com/shows/the-bachelor/cast"
    cast_page = requests.get(bach_cast_url)
    soup = BeautifulSoup(cast_page.content, "html.parser")

    LEAD_NAME_CELL = "B1"
    CONTESTANT_START_COL = "A"
    CONTESTANT_ROW = 7
    template = load_workbook(filename="Bach Template.xlsx")
    sheet = template.active

    cast = soup.find_all("div", class_="tile__name")

    lead_name = cast[0]
    sheet[LEAD_NAME_CELL] = lead_name.text
    host_name = cast[1]

    for el in cast[2:]:
        sheet[CONTESTANT_START_COL + str(CONTESTANT_ROW)] = el.text
        CONTESTANT_ROW += 1
    template.save(filename="Bach Template.xlsx")

def set_contestants():
    global CONTESTANT_LIST
    if cur_rose_ceremony.get() != "SELECT":
        print(cur_rose_ceremony.get()) #prints current rose ceremony
        new_contestants = []
        for c in range(0, len(contestant_vars)): #prints list of contestants selected
            if contestant_vars[c].get() == 1:
                new_contestants.append(CONTESTANT_LIST[c])
        
        with open('lib/contestant_list.pkl', 'wb') as f:
            pickle.dump(new_contestants, f)
        with open('lib/contestant_list.pkl', 'rb') as f:
            CONTESTANT_LIST = pickle.load(f)

def grade():
    file = load_workbook(filename="KeithD.xlsx")
    sheet = file.active
    rose_ceremony = ROSE_CEREMONY_META[cur_rose_ceremony.get()]
    sheet_contestants = []
    for i in range(rose_ceremony["starting_cell"], rose_ceremony["ending_cell"]+1):
        sheet_contestants.append(sheet[rose_ceremony["column"]+str(i)].value)
    wrong_contestants = set(CONTESTANT_LIST) - set(sheet_contestants)
    print(wrong_contestants)
    points = (rose_ceremony["num_contestants"] - len(wrong_contestants))*rose_ceremony["pp_correct_answer"]
    print(points)
    sheet[rose_ceremony["column"]+str(rose_ceremony["weekly_score_cell"])] = points
    file.save(filename="KeithD.xlsx")

if __name__ == "__main__":
    window = tk.Tk()
    window.title("The Bachelor/Bachelorette")

    #setup = tk.Button(text="Setup Season & Brackets")
    #setup.grid()

    tk.Label(text="Rose Ceremony Entry").grid()
    cur_rose_ceremony = tk.StringVar(window)
    cur_rose_ceremony.set("SELECT")
    rose_ceremony_selector = tk.OptionMenu(window, cur_rose_ceremony, *ROSE_CEREMONY_META.keys())
    rose_ceremony_selector.grid()

    checklist = tk.LabelFrame(window)
    global contestant_vars
    contestant_vars = []
    for contestant in CONTESTANT_LIST:
        var = tk.IntVar()
        contestant_vars.append(var)
        c = tk.Checkbutton(checklist, text=contestant, variable=var)
        c.grid(column=int(CONTESTANT_LIST.index(contestant)%3), row=int(CONTESTANT_LIST.index(contestant)/3))
    checklist.grid(columnspan=3)
    finalize_rc = tk.Button(text="Finalize Rose Ceremony", command=set_contestants)
    finalize_rc.grid()

    grade_brackets = tk.Button(text="Grade Brackets", command=grade)
    grade_brackets.grid()

    window.mainloop()

    
    